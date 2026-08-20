"""Build a queryable SQLite database from JRKJ data files 2-5."""

from __future__ import annotations

import os
import sqlite3
from pathlib import Path
from typing import Iterable, Iterator, Sequence

import pandas as pd
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DATA_DIR = PROJECT_ROOT / "data"
DEFAULT_DB = PROJECT_ROOT / "database" / "jrkj.sqlite3"
CHUNK_SIZE = 5_000

CSV_TABLES = {
    "financial_balance": DATA_DIR / "4" / "asharebalancesheet_202605261517.csv",
    "financial_cashflow": DATA_DIR / "4" / "asharecashflow_202605261518.csv",
    "financial_income": DATA_DIR / "4" / "ashareincome_202605261519.csv",
    "research_reports": DATA_DIR / "5" / "rr_main_202605281537.csv",
}

DATE_COLUMNS = {
    "ann_dt",
    "actual_ann_dt",
    "report_period",
    "s_holder_enddate",
    "collect_dt",
    "write_date",
    "publish_date",
}


def normalize_date(value: object) -> int | None:
    """Convert Excel, ISO, or numeric dates to an integer YYYYMMDD."""
    if value is None or pd.isna(value):
        return None
    if hasattr(value, "strftime"):
        return int(value.strftime("%Y%m%d"))
    digits = "".join(char for char in str(value).split(".")[0] if char.isdigit())
    if len(digits) >= 8:
        return int(digits[:8])
    return None


def normalize_windcode(sec_code: object, exchange_code: object) -> str | None:
    """Convert research-report security codes to Wind-style codes."""
    if sec_code is None or pd.isna(sec_code):
        return None
    code = str(sec_code).strip().split(".")[0].zfill(6)
    suffix = {
        "XSHG": "SH",
        "XSHE": "SZ",
        "XBSE": "BJ",
        "SH": "SH",
        "SZ": "SZ",
        "BJ": "BJ",
    }.get(str(exchange_code).strip().upper())
    return f"{code}.{suffix}" if suffix else code


def sqlite_value(value: object) -> object:
    """Convert pandas/numpy values into types supported by sqlite3."""
    if value is None or pd.isna(value):
        return None
    if hasattr(value, "item"):
        return value.item()
    return value


def batches(rows: Iterable[Sequence[object]], size: int) -> Iterator[list[Sequence[object]]]:
    batch: list[Sequence[object]] = []
    for row in rows:
        batch.append(row)
        if len(batch) == size:
            yield batch
            batch = []
    if batch:
        yield batch


def create_table_from_columns(
    connection: sqlite3.Connection, table: str, columns: Sequence[str]
) -> None:
    quoted = ", ".join(f'"{column}"' for column in columns)
    connection.execute(f'DROP TABLE IF EXISTS "{table}"')
    connection.execute(f'CREATE TABLE "{table}" ({quoted})')


def import_excel(
    connection: sqlite3.Connection, path: Path, table: str, sheet_name: str | None = None
) -> int:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    rows = sheet.iter_rows(values_only=True)
    columns = [str(value).strip().lower() for value in next(rows)]
    create_table_from_columns(connection, table, columns)

    placeholders = ", ".join("?" for _ in columns)
    insert_sql = f'INSERT INTO "{table}" VALUES ({placeholders})'

    def cleaned_rows() -> Iterator[tuple[object, ...]]:
        for row in rows:
            cleaned = []
            for column, value in zip(columns, row):
                cleaned.append(normalize_date(value) if column in DATE_COLUMNS else sqlite_value(value))
            yield tuple(cleaned)

    count = 0
    for batch in batches(cleaned_rows(), CHUNK_SIZE):
        connection.executemany(insert_sql, batch)
        count += len(batch)
    workbook.close()
    connection.commit()
    return count


def clean_csv_chunk(table: str, chunk: pd.DataFrame) -> pd.DataFrame:
    chunk.columns = [column.strip().lower() for column in chunk.columns]
    for column in DATE_COLUMNS.intersection(chunk.columns):
        chunk[column] = chunk[column].map(normalize_date).astype("Int64")
    if table == "research_reports":
        chunk.insert(
            0,
            "s_info_windcode",
            [
                normalize_windcode(code, exchange)
                for code, exchange in zip(chunk["sec_code"], chunk["exchange_code"])
            ],
        )
    return chunk


def import_csv(connection: sqlite3.Connection, path: Path, table: str) -> int:
    count = 0
    dtype = {"sec_code": "string"} if table == "research_reports" else None
    for index, chunk in enumerate(
        pd.read_csv(path, chunksize=CHUNK_SIZE, low_memory=False, dtype=dtype)
    ):
        clean_csv_chunk(table, chunk).to_sql(
            table,
            connection,
            if_exists="replace" if index == 0 else "append",
            index=False,
        )
        count += len(chunk)
    connection.commit()
    return count


def create_indexes(connection: sqlite3.Connection) -> None:
    statements = [
        "CREATE INDEX idx_shareholders_code_enddate ON shareholders(s_info_windcode, s_holder_enddate)",
        "CREATE INDEX idx_shareholders_code_anndate ON shareholders(s_info_windcode, ann_dt)",
        "CREATE INDEX idx_shareholders_name_code_date ON shareholders(s_holder_name, s_info_windcode, s_holder_enddate)",
        "CREATE INDEX idx_announcements_code_date ON announcements(s_info_windcode, ann_dt)",
        "CREATE INDEX idx_income_code_period ON financial_income(s_info_windcode, report_period)",
        "CREATE INDEX idx_income_code_anndate ON financial_income(s_info_windcode, ann_dt)",
        "CREATE INDEX idx_balance_code_period ON financial_balance(s_info_windcode, report_period)",
        "CREATE INDEX idx_balance_code_anndate ON financial_balance(s_info_windcode, ann_dt)",
        "CREATE INDEX idx_cashflow_code_period ON financial_cashflow(s_info_windcode, report_period)",
        "CREATE INDEX idx_cashflow_code_anndate ON financial_cashflow(s_info_windcode, ann_dt)",
        "CREATE INDEX idx_reports_code_date ON research_reports(s_info_windcode, publish_date)",
    ]
    for statement in statements:
        connection.execute(statement)
    connection.commit()


def build_database(output: Path, force: bool = False) -> dict[str, int]:
    output = output.resolve()
    if output.exists() and not force:
        raise FileExistsError(f"Database already exists: {output}. Use --force to rebuild it.")
    output.parent.mkdir(parents=True, exist_ok=True)
    temporary = output.with_suffix(output.suffix + ".tmp")
    temporary.unlink(missing_ok=True)

    counts: dict[str, int] = {}
    connection = sqlite3.connect(temporary)
    try:
        connection.execute("PRAGMA journal_mode = WAL")
        connection.execute("PRAGMA synchronous = NORMAL")
        counts["shareholders"] = import_excel(
            connection, DATA_DIR / "2" / "上市公司前十大股东.xlsx", "shareholders", "十大股东"
        )
        counts["announcements"] = import_excel(
            connection, DATA_DIR / "3" / "公司风险公告目录.xlsx", "announcements"
        )
        for table, path in CSV_TABLES.items():
            counts[table] = import_csv(connection, path, table)
        create_indexes(connection)
        connection.execute("PRAGMA optimize")
    except Exception:
        connection.close()
        temporary.unlink(missing_ok=True)
        raise
    else:
        connection.close()
        os.replace(temporary, output)
    return counts
